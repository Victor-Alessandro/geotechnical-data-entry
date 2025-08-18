import streamlit as st
import pandas as pd
import io
import os

from utils import ndc

from openpyxl import load_workbook
from typing import Optional
from utils.reusables import initialize_session_state

# Page configuration
st.set_page_config(
    page_title="Analyse Géotechnique et Fichiers",
    page_icon="🏗️",
    layout="wide",
)

# --- State Machine States ---
STATE_NONE_PROCESSED = "none_processed"
STATE_ONLY_GEO = "only_geo"
STATE_ONLY_DDC = "only_ddc"
STATE_ALL_PROCESSED = "all_processed"

# --- State Detection Functions ---
def get_current_state() -> str:
    """Determine current state based on session variables"""
    has_geo = st.session_state.get('final_geo_data') is not None
    has_ddc = st.session_state.get('final_ddc_data') is not None
    
    if has_geo and has_ddc:
        return STATE_ALL_PROCESSED
    elif has_geo:
        return STATE_ONLY_GEO
    elif has_ddc:
        return STATE_ONLY_DDC
    else:
        return STATE_NONE_PROCESSED

def display_state_message(state: str) -> None:
    """Display appropriate message based on current state"""
    if state == STATE_NONE_PROCESSED:
        st.warning("Ni le DDC ni l'étude de sol n'ont été traités")
    elif state == STATE_ONLY_GEO:
        st.info("Le DDC n'a pas été traité")
    elif state == STATE_ONLY_DDC:
        st.info("L'étude de sol n'a pas été traité")
    elif state == STATE_ALL_PROCESSED:
        st.success("Tous les documents ont été traités, veuillez trouver les documents de calcul pré-remplis")

# --- File Processing Functions ---
def get_base_excel_file(filename: str) -> bytes:
    """Load base Excel file from resources"""
    filepath = os.path.join("resources", "Excel", filename)
    if os.path.exists(filepath):
        with open(filepath, "rb") as f:
            return f.read()
    return None

def fill_hypotheses_sheet(workbook, geo_data: pd.DataFrame) -> None:
    """Fill the Hypothèses sheet with geo data from D23 to H23"""
    if 'Hypothèses' not in workbook.sheetnames:
        return
    
    sheet = workbook['Hypothèses']
    start_col = 4  # Column D
    start_row = 23
    
    # Fill up to 5 columns or until geo_data columns end
    max_cols = min(5, len(geo_data.columns))
    
    for col_idx in range(max_cols):
        if len(geo_data) > 0:  # Ensure there's at least one row
            cell = sheet.cell(row=start_row, column=start_col + col_idx)
            cell.value = geo_data.iloc[0, col_idx]

def fill_dimensionnement_sheet(workbook, ddc_data: pd.DataFrame) -> None:
    """Fill the Dimensionnement Général sheet with DDC data from B13 to D13+"""
    if 'Dimensionnement Général' not in workbook.sheetnames:
        return
    
    sheet = workbook['Dimensionnement Général']
    start_col = 3  # Column B
    start_row = 13
    
    # Fill up to 3 columns or until ddc_data columns end
    max_cols = min(3, len(ddc_data.columns))
    
    # Fill all rows of the dataframe
    for row_idx in range(len(ddc_data)):
        for col_idx in range(max_cols):
            cell = sheet.cell(row=start_row + row_idx, column=start_col + col_idx)
            cell.value = ddc_data.iloc[row_idx, col_idx]

def create_modified_excel(filename: str, geo_data: Optional[pd.DataFrame], ddc_data: Optional[pd.DataFrame]) -> bytes:
    """Create modified Excel file based on available data"""
    base_file = get_base_excel_file(filename)
    if not base_file:
        return None
    
    # Load workbook from bytes
    workbook = load_workbook(io.BytesIO(base_file))
    
    # Apply modifications based on available data
    if geo_data is not None and not geo_data.empty:
        fill_hypotheses_sheet(workbook, geo_data)
    
    # Only fill DDC data for Portance files and if DDC has <= 3 columns
    if (ddc_data is not None and not ddc_data.empty and 
        len(ddc_data.columns) <= 3 and 
        filename in ['Portance v1.xlsm', 'Portance v2.xlsm']):
        fill_dimensionnement_sheet(workbook, ddc_data)
    
    # Save to bytes
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()

def get_download_filename(base_filename: str, state: str) -> str:
    """Generate appropriate filename based on state"""
    name, ext = os.path.splitext(base_filename)
    if state in [STATE_ONLY_GEO, STATE_ONLY_DDC, STATE_ALL_PROCESSED]:
        return f"{name}_modifie{ext}"
    return base_filename

# --- UI Display Functions ---
def display_excel_downloads(state: str) -> None:
    """Display Excel file generation and download buttons based on current state"""
    st.header("📊 Fichiers Excel de Calcul")
    
    excel_files = ["Pieuhor.xlsm", "Portance v1.xlsm", "Portance v2.xlsm"]
    
    if state != STATE_NONE_PROCESSED:     
        if st.button("🔧 Générer les Fichiers Excel", key="generate_excel_button", help="Cliquez pour générer tous les fichiers Excel avec les données traitées."):
            with st.spinner("Génération des fichiers Excel en cours..."):
                try:
                    geo_data = st.session_state.get('final_geo_data')
                    ddc_data = st.session_state.get('final_ddc_data')
                    
                    # Generate all Excel files
                    generated_files = {}
                    for filename in excel_files:
                        modified_file = create_modified_excel(filename, geo_data, ddc_data)
                        if modified_file:
                            download_filename = get_download_filename(filename, state)
                            generated_files[filename] = {
                                'data': modified_file,
                                'filename': download_filename
                            }
                        else:
                            st.error(f"Erreur lors de la génération de {filename}")
                    
                    if generated_files:
                        st.session_state.generated_excel_files = generated_files
                        st.toast("Fichiers Excel générés avec succès!", icon="🎉")
                    else:
                        st.session_state.generated_excel_files = None
                        st.error("Erreur: Aucun fichier Excel n'a pu être généré.")
                        
                except Exception as e:
                    st.session_state.generated_excel_files = None
                    st.error(f"Erreur critique lors de la génération: {str(e)}")       
   
    # Show download buttons
    col1, col2, col3 = st.columns(3)
    columns = [col1, col2, col3]
    
    for idx, filename in enumerate(excel_files):
        with columns[idx]:
            st.subheader(filename.replace('.xlsm', ''))
            
            if state == STATE_NONE_PROCESSED:
                # Original files - direct download
                base_file = get_base_excel_file(filename)
                if base_file:
                    st.download_button(
                        label=f"📥 Télécharger {filename}",
                        data=base_file,
                        file_name=filename,
                        mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                        key=f"download_original_{idx}"
                    )
                else:
                    st.error(f"Fichier {filename} non trouvé")
            else:
                # Modified files - download after generation
                generated_files = st.session_state.get('generated_excel_files', {})
                if generated_files and filename in generated_files:
                    file_info = generated_files[filename]
                    st.download_button(
                        label=f"📥 Télécharger {file_info['filename']}",
                        data=file_info['data'],
                        file_name=file_info['filename'],
                        mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                        key=f"download_modified_{idx}"
                    )
                else:
                    st.info("Veuillez d'abord générer les fichiers Excel")

def display_dataframe_section(title: str, data: Optional[pd.DataFrame]) -> None:
    """Display dataframe section with title"""
    st.subheader(title)
    if data is not None and not data.empty:
        st.dataframe(data, use_container_width=True)
        st.write(f"**Colonnes:** {', '.join(data.columns.tolist())}")
    else:
        st.info(f"{title.lower()} indisponibles")

def display_data_summary() -> None:
    """Display summary of processed data"""
    st.divider()
    st.header("📋 Résumé des Données Traitées")
    
    col1, col2 = st.columns(2)
    
    with col1:
        display_dataframe_section("Données DDC", st.session_state.get('final_ddc_data'))
    
    with col2:
        display_dataframe_section("Données Géotechniques", st.session_state.get('final_geo_data'))

def display_ndc_download():
    # --- Section for Note de Calcul (NDC) ---
    st.divider()
    st.header("📝 Note de Calcul")

    if st.session_state.get('calculs_excel') and st.session_state.get('plan_masse'):
        st.success("Les fichiers requis (Calcul de portance et Plan de masse) sont chargés.")

        st.subheader("Génération du Document Final")
        
        if st.button("🔧 Générer la Note de Calcul", key="generate_doc_button_fichiers", help="Cliquez pour lancer la génération du document Word."):
            validation_errors = ndc.FormValidator.validate_required_fields(st.session_state)
            if validation_errors:
                for error in validation_errors:
                    st.error(error)
                st.session_state.generated_docx = None
            else:
                with st.spinner("Génération du document en cours... Veuillez patienter."):
                    try:
                        # after the document has been generated I should clear at least the files from the respective
                        # pages. At least related with the diameters table there is a side effect, and its unlikely to be
                        # the only one.
                        
                        template_path ="./resources/"                        
                        if st.session_state.horiz_excel:
                            if st.session_state.diameter_option == "Plusieurs diamètres":
                                template_path = template_path + "aciers2.docx"
                            else:
                                template_path = template_path + "aciers.docx"
                        else:
                            if st.session_state.diameter_option == "Plusieurs diamètres":
                                template_path = template_path + "normal2.docx"
                            else:
                                template_path = template_path + "normal.docx"
                         
                        doc_generator = ndc.DocumentGenerator( template_path )
                        context = ndc.build_document_context(st.session_state)
                        
                        st.session_state.highlighted_image = ndc.generate_highlighted_seismic_table_image_data(
                            st.session_state.sismicite,
                            st.session_state.type_batiment
                        )
                        st.session_state.generated_image_paths = doc_generator.prepare_images(st.session_state) # For download buttons

                        # Prepare images and dataframes for display and inclusion
                        st.session_state.generated_image_paths = doc_generator.prepare_images(st.session_state)

                        # Generate the document
                        docx_file = doc_generator.generate_document(context)
                        doc_generator.cleanup()

                        if docx_file:
                            st.session_state.generated_docx = docx_file
                            st.toast("Document généré avec succès!", icon="🎉")
                        else:
                            st.session_state.generated_docx = None
                            st.error("Erreur: Le document n'a pas pu être généré.")
                    except Exception as e:
                        st.session_state.generated_docx = None
                        st.error(f"Erreur critique lors de la génération: {str(e)}")

        if st.session_state.generated_docx:
            download_file_name = f"NDC_{st.session_state.get('affaire', '00000')}_{st.session_state.get('commune', 'Commune').split(' (')[0].replace(' ','_')}.docx"
            st.download_button(
                label="📥 Télécharger la Note de Calcul (.docx)",
                data=st.session_state.generated_docx,
                file_name=download_file_name,
                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                key="download_docx_btn_fichiers"
            )

    else:
        st.info("Il faut d'abord que le calcul de la portance et une image du plan de masse soient chargés.")


# --- Main Application Logic ---
def initialize_required_session_state() -> None:
    """Initialize all required session state variables"""
    KEYS_TO_INIT = {
        'final_geo_data': None,
        'final_ddc_data': None,
        'generated_docx': None,
        'generated_excel_files': None,
        'calculs_excel': None,
        'plan_masse': None,
        'generated_image_paths': None,
        'display_dataframes': None,
    }
    initialize_session_state(KEYS_TO_INIT)

def main():
    """Main application function implementing finite state machine"""

    initialize_required_session_state()
    
    st.title("🔽 Téléchargement des Fichiers")
    st.write("Utilisez les pages dédiées pour traiter vos documents, puis revenez ici pour générer et télécharger les fichiers finaux.")
    
    current_state = get_current_state()
    
    display_state_message(current_state)
    
    st.divider()
    
    display_excel_downloads(current_state)
    
    display_data_summary()
    
    display_ndc_download()
    
if __name__ == "__main__":
    main()
